import time
import json
import threading
from urllib.parse import urlparse
from datetime import datetime
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import streamlit as st

def setup_driver():
    options = Options()
    options.add_argument('--disable-popup-blocking')
    options.add_argument('--disable-notifications')
    options.add_argument('--disable-infobars')
    driver = webdriver.Chrome(options=options)
    driver.maximize_window()
    return driver

class TestCaseManager:
    def __init__(self, website_name):
        self.website_name = website_name
        self.testcases_filepath = f"{website_name}_testcases.json"
        self.login_performed = False  # Tracks if login was already performed

    def load_test_cases(self):
        try:
            with open(self.testcases_filepath, "r") as file:
                return json.load(file)
        except FileNotFoundError:
            return {"login_credentials": {}, "test_cases": []}

    def save_test_cases(self, data):
        with open(self.testcases_filepath, "w") as file:
            json.dump(data, file, indent=4)

    def scroll_page(self, driver):
        print("Scrolling down and up slowly...")
        for _ in range(5):
            driver.execute_script("window.scrollBy(0, window.innerHeight / 5);")
            time.sleep(1)

        for _ in range(5):
            driver.execute_script("window.scrollBy(0, -window.innerHeight / 5);")
            time.sleep(1)

    def generate_test_cases(self, url, driver, login_credentials=None):
        print("Opening URL...")
        driver.get(url)

        data = self.load_test_cases()
        if login_credentials and not self.login_performed:
            data["login_credentials"] = {
                "email": login_credentials[0],
                "password": login_credentials[1]
            }
            print("Logging in with provided credentials...")
            self.perform_login(driver, login_credentials)
            self.login_performed = True

        WebDriverWait(driver, 20).until(
            lambda d: d.execute_script("return document.readyState") == "complete"
        )

        if not any(tc["action"] == "scroll" for tc in data["test_cases"]):
            data["test_cases"].append({
                "url": url,
                "action": "scroll",
                "expected_output": "Page scrolled up and down successfully"
            })
            print("Generated test case for scrolling.")

        clickable_elements = driver.find_elements(By.XPATH, "//*[@onclick or self::button or self::a]")

        for element in clickable_elements:
            try:
                xpath = self.generate_xpath(element, driver)
                tag_name = element.tag_name
                element_text = element.text.strip() or f"Unnamed {tag_name.capitalize()}"
                if not any(tc.get("element_xpath") == xpath for tc in data["test_cases"]):
                    data["test_cases"].append({
                        "url": url,
                        "element_xpath": xpath,
                        "action": "click",
                        "expected_output": f"Click action performed on {element_text} ({tag_name})"
                    })
                    print(f"Generated test case for {tag_name}: '{element_text}', XPath: {xpath}")
            except Exception as e:
                print(f"Error generating test case for element: {e}")

        self.save_test_cases(data)
        print("Test cases generated and saved.")

    def perform_login(self, driver, login_credentials):
        try:
            email_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='email' or @name='email']"))
            )
            password_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='password' or @name='password']"))
            )
            email_field.send_keys(login_credentials[0])
            password_field.send_keys(login_credentials[1])
            password_field.send_keys(Keys.RETURN)
            print("Login successful.")
        except Exception as e:
            print(f"Error during login: {e}")

    def generate_xpath(self, element, driver):
        try:
            return driver.execute_script(
                """
                var getXPath = function(element) {
                    if (element.id !== '') {
                        return 'id("' + element.id + '")';
                    }
                    var path = [];
                    while (element.nodeType === Node.ELEMENT_NODE) {
                        var index = 0;
                        var sibling = element.previousSibling;
                        while (sibling) {
                            if (sibling.nodeType === Node.ELEMENT_NODE && sibling.tagName === element.tagName) {
                                index++;
                            }
                            sibling = sibling.previousSibling;
                        }
                        var tagName = element.tagName.toLowerCase();
                        var part = tagName + (index ? '[' + (index + 1) + ']' : '');
                        path.unshift(part);
                        element = element.parentNode;
                    }
                    return '/' + path.join('/');
                };
                return getXPath(arguments[0]);
                """, element
            )
        except Exception as e:
            raise ValueError(f"Unable to generate XPath for element: {e}")


class TestRunner:
    def __init__(self, website_name):
        self.testcases_filepath = f"{website_name}_testcases.json"
        self.report_filepath = f"{website_name}_test_reports.xlsx"

    def load_test_cases(self):
        try:
            with open(self.testcases_filepath, "r") as file:
                return json.load(file)
        except FileNotFoundError:
            return {"login_credentials": {}, "test_cases": []}

    def setup_excel_report(self):
        try:
            workbook = load_workbook(self.report_filepath)
        except FileNotFoundError:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Test Report"
            headers = ["Date", "Time", "Element XPath", "Action", "Status"]
            for col, header in enumerate(headers, start=1):
                sheet[f"{get_column_letter(col)}1"] = header
            workbook.save(self.report_filepath)
        return workbook

    def log_result(self, action, element_xpath=None, status="Pass"):
        workbook = self.setup_excel_report()
        sheet = workbook.active
        date_time = datetime.now()
        new_row = [date_time.strftime("%Y-%m-%d"), date_time.strftime("%H:%M:%S"), element_xpath or "N/A", action, status]
        sheet.append(new_row)
        workbook.save(self.report_filepath)

    def run_tests(self):
        data = self.load_test_cases()
        login_credentials = data.get("login_credentials", {})
        test_cases = data.get("test_cases", [])

        if not test_cases:
            print("No test cases to run.")
            return

        driver = setup_driver()

        if login_credentials:
            self.perform_login(driver, test_cases[0]["url"], login_credentials)

        current_url = ""
        for case in test_cases:
            try:
                if case["url"] != current_url:
                    driver.get(case["url"])
                    current_url = case["url"]

                if case["action"] == "scroll":
                    self.perform_scroll(driver)
                else:
                    element = driver.find_element(By.XPATH, case["element_xpath"])
                    if case["action"] == "click":
                        element.click()
                        
                        WebDriverWait(driver, 5).until(
                            lambda d: d.current_url != current_url
                        )
                        print("Page changed after click, navigating back.")
                        driver.back()

                self.log_result(case["action"], case.get("element_xpath"), "Pass")
            except Exception as e:
                print(f"Error with element {case.get('element_xpath')}: {e}")
                self.log_result(case["action"], case.get("element_xpath"), "Fail")

        driver.quit()
        print("Test execution complete. Check the Excel report.")

    def perform_scroll(self, driver):
        print("Performing scroll test...")
        for _ in range(5):
            driver.execute_script("window.scrollBy(0, window.innerHeight / 5);")
            time.sleep(1)
        for _ in range(5):
            driver.execute_script("window.scrollBy(0, -window.innerHeight / 5);")
            time.sleep(1)

    def perform_login(self, driver, url, login_credentials):
        try:
            driver.get(url)
            email_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='email' or @name='email']"))
            )
            password_field = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, "//input[@type='password' or @name='password']"))
            )
            email_field.send_keys(login_credentials["email"])
            password_field.send_keys(login_credentials["password"])
            password_field.send_keys(Keys.RETURN)
            print("Login successful.")
        except Exception as e:
            print(f"Error during login: {e}")

    def schedule_tests(self, interval):
        def run_periodically():
            while True:
                self.run_tests()
                time.sleep(interval)

        threading.Thread(target=run_periodically, daemon=True).start()
        print(f"Tests scheduled every {interval} seconds. Press Ctrl+C to stop.")


def get_site_name_from_url(url):
    return urlparse(url).netloc.split('.')[0]

def main():
    st.set_page_config(page_title="Autotest Pro", layout="centered")
    st.title("AUTOTEST PRO")
    st.markdown(
        """
        <h4 style="color: grey; text-align: center;">
            EMPOWERING SEAMLESS AUTOMATION FOR FLAWLESS PERFORMANCE
        </h4>
        """,
        unsafe_allow_html=True,
    )
    st.markdown("---")

    choice = st.sidebar.radio(
        "Menu",  
        ["Generate Test Cases", "Run Test Cases", "Schedule Automated Testing", "Exit"],
        index=0,
        key="menu_radio",
    )

    if choice == "Generate Test Cases":
        st.header("Generate Test Cases")
        url = st.text_input("Enter the URL of the website:")
        website_name = st.text_input("Enter a name for the website (used for file storage):")
        require_login = st.radio("Does the website require login?", ["No", "Yes"])
        login_credentials = None

        if require_login == "Yes":
            email = st.text_input("Enter email:")
            password = st.text_input("Enter password:", type="password")
            login_credentials = (email, password)

        if st.button("Generate"):
            if url and website_name:
                driver = setup_driver()
                manager = TestCaseManager(website_name)
                manager.generate_test_cases(url, driver, login_credentials)
                driver.quit()
            else:
                st.error("Please provide all required details.")

    elif choice == "Run Test Cases":
        st.header("Run Test Cases")
        website_name = st.text_input("Enter the name of the website (used for file storage):")

        if st.button("Run"):
            if website_name:
                runner = TestRunner(website_name)
                runner.run_tests()
            else:
                st.error("Please provide the website name.")

    elif choice == "Schedule Automated Testing":
        st.header("Schedule Automated Testing")
        website_name = st.text_input("Enter the name of the website:")
        interval = st.number_input("Enter the interval in seconds for scheduled testing:", min_value=1, step=1)

        if st.button("Schedule"):
            if website_name and interval > 0:
                runner = TestRunner(website_name)
                runner.schedule_tests(interval)
            else:
                st.error("Please provide valid inputs.")

    elif choice == "Exit":
        st.header("Thank you for using the tool!")

    st.markdown(
        """
        <style>
            /* App Background */
            .stApp {
                background-color: white;
                background-image: linear-gradient(to bottom right, transparent, transparent 50%, #dcdcdc 50%);
            }

            /* Header Text */
            h1 {
                color: black;
                text-align: center;
            }
            
            h2,h3,h5,h6 {
                color: darkgrey;
                text-align: center;
            }

            /* Subheading Tagline */
            h4 {
                font-size: 18px;
                font-style: italic;
                color: grey;
            }

            /* Input Fields */
            input, textarea, select {
                background-color: #d3d3d3 !important; /* Light grey */
                color: black;
                border: 1px solid #ccc;
                border-radius: 5px;
                padding: 8px;
            }

            /* Input Labels and Sidebar Options */
            label, .stRadio label, .stSidebar label {
                color: #4f4f4f !important; /* Dark grey */
                font-weight: bold;
            }

            /* Sidebar Styling */
            [data-testid="stSidebar"] {
                background-color: #909495;
                border-right: 1px solid #ddd;
            }

            /* Sidebar Option Text (Home Screen Options) */
            .stRadio div[data-baseweb="block"] {
                color: black !important; /* Dark grey text for main options */
                font-weight: bold;
            }

            /* Buttons for Radio Options (Yes/No) */
            div[data-baseweb="radio"] > div {
                background-color: #dcdcdc !important; /* Light grey background for buttons */
                border-radius: 5px;
                color: black !important; /* Dark grey text */
                font-weight: bold;
                margin-right: 10px;
                padding: 10px;
                display: inline-block;
                text-align: center;
                cursor: pointer;
            }

            /* Highlight selected radio button */
            div[data-baseweb="radio"] > div:hover {
                background-color: black; /* Slightly darker grey */
            }
            div[data-baseweb="radio"] > div[data-checked="true"] {
                background-color: black; /* Slightly darker grey */
                border: 2px solid #696969; /* Darker border */
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()
